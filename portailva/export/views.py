import string

import unicodedata
from django.http import HttpResponse
from django.views.generic import TemplateView

from portailva.association.models import Association, Requirement
from portailva.export.mixins import AbleToExportMixin


class ExportView(AbleToExportMixin, TemplateView):
    template_name = 'export/export.html'

    def post(self, request):
        return self.export_xlsx(request.POST['filter'], request.POST.getlist('data'))

    @staticmethod
    def convertToTitle(num):
        title = ''
        alist = string.ascii_uppercase
        while num:
            mod = (num - 1) % 26
            num = int((num - mod) / 26)
            title += alist[mod]
        return title[::-1]

    def export_xlsx(self, category='ALL', datas=('BASIC', 'VALIDATIONS', 'PRESIDENT', 'BANK')):

        # Create the Excel file
        import openpyxl
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=associations.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.get_active_sheet()
        ws.title = "Associations"

        row_num = 0

        # Define columns to export
        columns = [ExportColumn('Nom', 'name', 30)]
        if 'BASIC' in datas:
            columns.append(ExportColumn('Acronyme', 'acronym'))
            columns.append(ExportColumn('Categorie', 'category.name', 22))
            if category == 'ALL':
                columns.append(ExportColumn('Active', 'is_active'))
            columns.append(ExportColumn('Validé', 'is_validated'))
        if 'VALIDATIONS' in datas:
            for requirement in Requirement.objects.get_all_active():
                columns.append(RequirementExportColumn(requirement))
        if 'PRESIDENT' in datas:
            columns.append(ExportColumn('Président (Nom)', 'mandates.last().peoples.first().__str__()'))
            columns.append(ExportColumn('Président (Téléphone)', 'mandates.last().peoples.first().phone'))
            columns.append(ExportColumn('Président (Email)', 'mandates.last().peoples.first().email'))
        if 'BANK' in datas:
            columns.append(ExportColumn('IBAN', 'iban'))
            columns.append(ExportColumn('BIC', 'bic'))
        if 'BNP' in datas:
            columns = [ExportColumn('TIERS - NOM OU RAISON SOCIALE', 'name', options=['caps', 'accents']),
                       ExportColumn('TIERS - TYPE', value='Autre'),
                       ExportColumn('TIERS - REFERENCE INTERNE', 'id', prefix='CVA-'),
                       ExportColumn('TIERS - PAYS DE RESIDENCE', value='FRANCE'),
                       ExportColumn('TIERS - ADRESSE (NUMERO, VOIE)'),
                       ExportColumn('TIERS - COMPLEMENT ADRESSE'),
                       ExportColumn('TIERS - CODE POSTAL'),
                       ExportColumn('TIERS - VILLE'),
                       ExportColumn('TIERS - TELEPHONE'),
                       ExportColumn('TIERS - MOBILE'),
                       ExportColumn('TIERS - FAX'),
                       ExportColumn('TIERS - EMAIL'),
                       ExportColumn('TIERS - SIRET'),
                       ExportColumn('COMPTE - CODE SWIFT OU BIC', 'bic'),
                       ExportColumn('COMPTE - NUMERO IBAN', 'iban'),
                       ExportColumn('COMPTE - NUMERO BBAN'),
                       ExportColumn('COMPTE - NUMERO RIB'),
                       ExportColumn('COMPTE - CODE LOCAL (ABA OU ROUTING NUMBER)'),
                       ExportColumn('COMPTE - NOM DE LA BANQUE'),
                       ExportColumn('COMPTE - ADRESSE DE LA BANQUE'),
                       ExportColumn('COMPTE - COMPLEMENT ADRESSE DE LA BANQUE'),
                       ExportColumn('COMPTE - VILLE DE LA BANQUE'),
                       ExportColumn('COMPTE - PAYS DE LA BANQUE', value='FRANCE'),
                       ExportColumn('MANDAT- SDD MIGRE'),
                       ExportColumn('MANDAT- RUM'),
                       ExportColumn('MANDAT- SCHEMA'),
                       ExportColumn('MANDAT- TYPE'),
                       ExportColumn('MANDAT- DATE DE SIGNATURE')]

        # Create first line of document
        for col_num in range(len(columns)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = columns[col_num].name
            ws.column_dimensions[ExportView.convertToTitle(col_num+1)].width = columns[col_num].column_size

        # Fetch the set of data
        default_set = Association.objects.all().order_by('name')
        if category == 'ALIVE':
            default_set = default_set.filter(is_active=True)
        elif category == 'DEAD':
            default_set = default_set.filter(is_active=False)

        # Write data for each element in the set
        for obj in default_set:
            row_num += 1
            for col_num in range(len(columns)):
                c = ws.cell(row=row_num + 1, column=col_num + 1)
                c.value = columns[col_num].value_for(obj)

        wb.save(response)
        return response


class ExportColumn(object):

    def __init__(self, name, prop='', column_size=17, prefix='', value='', options=None):
        if options is None:
            options = []

        self.column_size = column_size
        self.name = name
        self.prop = prop
        self.prefix = prefix
        self.options = options
        self.value = value

    def value_for(self, association):
        value = str(self.prefix)
        try:
            if self.prop:
                val = str(eval("association."+self.prop))
            else:
                val = str(self.value)

            if 'accents' in self.options:
                val = ''.join((c for c in unicodedata.normalize('NFD', val) if unicodedata.category(c) != 'Mn'))
            if 'caps' in self.options:
                val = val.upper()

            return value + val

        except AttributeError:
            return ''


class RequirementExportColumn(ExportColumn):
    def __init__(self, requirement, **kwargs):
        super().__init__(requirement.name, **kwargs)
        self.requirement = requirement

    def value_for(self, association):
        return self.requirement.is_achieved(association.id)

